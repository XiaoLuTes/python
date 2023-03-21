import openpyxl

class Excel:

    def excel_read(self, filename):
        # sheet页面，row行，column列，cell格子
        filename2 = openpyxl.load_workbook(filename)
        sheet_names = filename2.sheetnames
        for sheet in sheet_names:
            sheet2 = filename2['' + sheet + '']
            max_row = sheet2.max_row
            # 记录最大行
            max_column = sheet2.max_column
            # 记录最大列
            datas = []
            for row in range(1, max_row + 1):
                data_list = []
                for column in range(1, max_column + 1):
                    data = sheet2.cell(row=row, column=column).value
                    data_list.append(data)
                datas.append(data_list)
                # # 通过定位sheet里的位置row，column提取cell内的值
                # # 并写入data_list
            title = datas[0]
            last_data = []
            for value in datas[1:]:
                new_data = dict(zip(title, value))
                last_data.append(new_data)
            print(last_data)
            return last_data

    def excel_write(self, filename, row, column, result):
        # 写入excel
        # data = Excel().excel_read(filename)
        file = openpyxl.load_workbook(filename)
        sheet_names = file.sheetnames
        for sheet in sheet_names:
            sheet2 = file['' + sheet + '']
            sheet2.cell(row=row, column=column).value = result
            file.save(filename)
